from __future__ import annotations

import hashlib
import re
from pathlib import Path
from typing import Any, Protocol

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field
from pypdf import PdfReader

from procureops.domain.procurement import ProcurementLine


class IntakeEvidence(BaseModel):
    model_config = ConfigDict(frozen=True)

    field_name: str
    line_number: int | None = None
    source_type: str
    source_id: str
    locator: str
    confidence: float = Field(ge=0, le=1)


class IntakeResult(BaseModel):
    model_config = ConfigDict(frozen=True)

    artifact_id: str
    source_type: str
    source_sha256: str
    lines: tuple[ProcurementLine, ...]
    evidence: tuple[IntakeEvidence, ...]
    questions: tuple[str, ...] = ()


class VisionExtractor(Protocol):
    def extract(self, path: Path) -> list[dict[str, Any]]: ...


class TextExtractor(Protocol):
    def extract(self, text: str) -> list[dict[str, Any]]: ...


class FakeVisionExtractor:
    def __init__(self, scripted_lines: list[dict[str, Any]]) -> None:
        self.scripted_lines = scripted_lines
        self.calls: list[Path] = []

    def extract(self, path: Path) -> list[dict[str, Any]]:
        self.calls.append(path)
        return self.scripted_lines


HEADER_ALIASES = {
    "description": {"description", "品名", "物料描述", "配件名称", "名称"},
    "quantity": {"quantity", "qty", "数量"},
    "unit": {"unit", "单位"},
    "part_number": {"part_number", "sku", "零件号", "物料编码"},
    "equipment_model": {"equipment_model", "设备型号", "机型"},
}


class IntakeService:
    def __init__(
        self,
        *,
        vision_extractor: VisionExtractor | None = None,
        text_extractor: TextExtractor | None = None,
    ) -> None:
        self.vision_extractor = vision_extractor
        self.text_extractor = text_extractor

    def from_text(self, text: str, *, artifact_id: str = "inline-text") -> IntakeResult:
        if not text.strip():
            return self._result(
                artifact_id=artifact_id,
                source_type="text",
                raw=b"",
                rows=[],
                locators=[],
                questions=("请提供至少一条采购明细。",),
            )
        rows: list[dict[str, Any]] = []
        locators: list[str] = []
        for line_number, raw_line in enumerate(text.splitlines(), start=1):
            stripped = raw_line.strip()
            if not stripped:
                continue
            parsed = self._parse_text_line(stripped)
            if parsed is not None:
                rows.append(parsed)
                locators.append(f"line:{line_number}")
        if not rows and self.text_extractor is not None:
            rows = self.text_extractor.extract(text)
            locators = [f"model:line:{index}" for index in range(1, len(rows) + 1)]
        questions = () if rows else ("未识别采购行，请补充零件号、名称、数量和单位。",)
        return self._result(
            artifact_id=artifact_id,
            source_type="text",
            raw=text.encode("utf-8"),
            rows=rows,
            locators=locators,
            questions=questions,
        )

    def from_file(self, path: Path) -> IntakeResult:
        if not path.is_file():
            raise FileNotFoundError(path)
        suffix = path.suffix.casefold()
        if suffix in {".txt", ".md", ".csv"}:
            return self.from_text(
                path.read_text(encoding="utf-8-sig"),
                artifact_id=path.name,
            )
        if suffix == ".pdf":
            return self._from_pdf(path)
        if suffix in {".xlsx", ".xlsm"}:
            return self._from_excel(path)
        if suffix in {".png", ".jpg", ".jpeg", ".webp"}:
            return self._from_image(path)
        raise ValueError(f"unsupported intake format: {suffix}")

    @staticmethod
    def _parse_text_line(text: str) -> dict[str, Any] | None:
        parts = [part.strip() for part in re.split(r"[|,，\t]", text)]
        if len(parts) >= 4:
            return {
                "part_number": parts[0] or None,
                "description": parts[1],
                "quantity": parts[2],
                "unit": parts[3],
                "equipment_model": parts[4] if len(parts) > 4 else None,
            }
        sku_match = re.search(r"\b(DEMO-[A-Z0-9-]+)\b", text, re.IGNORECASE)
        quantity_match = re.search(
            r"(?:[xX×*]|数量[:：]?\s*)(\d+(?:\.\d+)?)\s*([\u4e00-\u9fffA-Za-z]+)?",
            text,
        )
        if sku_match and quantity_match:
            description = text[: sku_match.start()] + text[sku_match.end() : quantity_match.start()]
            description = description.strip(" -:：，,") or sku_match.group(1)
            return {
                "part_number": sku_match.group(1).upper(),
                "description": description,
                "quantity": quantity_match.group(1),
                "unit": quantity_match.group(2) or "件",
            }
        return None

    def _from_pdf(self, path: Path) -> IntakeResult:
        reader = PdfReader(path)
        page_texts = [page.extract_text() or "" for page in reader.pages]
        combined = "\n".join(page_texts)
        result = self.from_text(combined, artifact_id=path.name)
        raw = path.read_bytes()
        if not result.lines and self.vision_extractor is not None:
            rows = self.vision_extractor.extract(path)
            return self._result(
                artifact_id=path.name,
                source_type="pdf_vision",
                raw=raw,
                rows=rows,
                locators=[f"pdf:region:{index}" for index in range(1, len(rows) + 1)],
                questions=() if rows else result.questions,
            )
        return result.model_copy(
            update={
                "source_type": "pdf",
                "source_sha256": hashlib.sha256(raw).hexdigest(),
            }
        )

    def _from_excel(self, path: Path) -> IntakeResult:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            rows: list[dict[str, Any]] = []
            locators: list[str] = []
        else:
            header_index = None
            headers: dict[str, int] = {}
            for candidate_index, candidate_row in enumerate(values[:50]):
                candidate_headers = self._map_headers(candidate_row)
                if candidate_headers:
                    header_index = candidate_index
                    headers = candidate_headers
                    break
            rows = []
            locators = []
            if header_index is not None:
                for row_index, values_row in enumerate(
                    values[header_index + 1 :], start=header_index + 2
                ):
                    row = {
                        field: values_row[index] if index < len(values_row) else None
                        for field, index in headers.items()
                    }
                    if row.get("description") and row.get("quantity"):
                        rows.append(row)
                        locators.append(f"{sheet.title}!A{row_index}")
        workbook.close()
        questions = () if rows else ("Excel 中未找到包含品名和数量的采购行。",)
        return self._result(
            artifact_id=path.name,
            source_type="excel",
            raw=path.read_bytes(),
            rows=rows,
            locators=locators,
            questions=questions,
        )

    def _from_image(self, path: Path) -> IntakeResult:
        if self.vision_extractor is None:
            return self._result(
                artifact_id=path.name,
                source_type="image",
                raw=path.read_bytes(),
                rows=[],
                locators=[],
                questions=("图片需要视觉模型提取，请配置 VisionExtractor。",),
            )
        rows = self.vision_extractor.extract(path)
        return self._result(
            artifact_id=path.name,
            source_type="image",
            raw=path.read_bytes(),
            rows=rows,
            locators=[f"image:region:{index}" for index in range(1, len(rows) + 1)],
            questions=() if rows else ("图片中未识别到采购明细。",),
        )

    @staticmethod
    def _map_headers(values: tuple[Any, ...]) -> dict[str, int]:
        headers: dict[str, int] = {}
        for index, value in enumerate(values):
            normalized = str(value or "").strip().casefold()
            for field, aliases in HEADER_ALIASES.items():
                if normalized in {alias.casefold() for alias in aliases}:
                    headers[field] = index
        required = {"description", "quantity"}
        if not required.issubset(headers):
            return {}
        return headers

    @staticmethod
    def _result(
        *,
        artifact_id: str,
        source_type: str,
        raw: bytes,
        rows: list[dict[str, Any]],
        locators: list[str],
        questions: tuple[str, ...],
    ) -> IntakeResult:
        lines: list[ProcurementLine] = []
        evidence: list[IntakeEvidence] = []
        for index, row in enumerate(rows, start=1):
            unit = str(row.get("unit") or "件").strip()
            line = ProcurementLine(
                line_number=index,
                description=str(row.get("description") or "").strip(),
                quantity=row.get("quantity"),
                unit=unit,
                part_number=(str(row["part_number"]).strip() if row.get("part_number") else None),
                equipment_model=(
                    str(row["equipment_model"]).strip() if row.get("equipment_model") else None
                ),
            )
            lines.append(line)
            locator = locators[index - 1]
            for field in ("description", "quantity", "unit", "part_number"):
                if getattr(line, field) is not None:
                    evidence.append(
                        IntakeEvidence(
                            field_name=field,
                            line_number=index,
                            source_type=source_type,
                            source_id=artifact_id,
                            locator=locator,
                            confidence=1.0 if source_type in {"text", "excel"} else 0.85,
                        )
                    )
        return IntakeResult(
            artifact_id=artifact_id,
            source_type=source_type,
            source_sha256=hashlib.sha256(raw).hexdigest(),
            lines=tuple(lines),
            evidence=tuple(evidence),
            questions=questions,
        )
